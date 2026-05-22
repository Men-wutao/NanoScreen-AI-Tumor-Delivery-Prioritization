"""Build Supplementary Excel File C from repository-relative screening outputs."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path(__file__).resolve().parents[1]
RESULTS_DIR = PROJECT_ROOT / "results"
SCREENING_DIR = RESULTS_DIR / "formula_generation"
SUPPLEMENTARY_DIR = PROJECT_ROOT / "supplementary"
OUTPUT_FILE = SUPPLEMENTARY_DIR / "Supplementary_Excel_File_C_virtual_screening_outputs.xlsx"


def value_to_text(value: Any) -> Any:
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def item_value_json(path: Path) -> pd.DataFrame:
    with path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    return pd.DataFrame([{"item": key, "value": value_to_text(value)} for key, value in payload.items()])


def read_optional_csv(file_name: str) -> pd.DataFrame:
    path = SCREENING_DIR / file_name
    return pd.read_csv(path) if path.exists() else pd.DataFrame()


def screening_summary(meta: dict[str, Any]) -> pd.DataFrame:
    filter_report = meta.get("filter_report", {}) if isinstance(meta.get("filter_report", {}), dict) else {}
    rows = [
        ("candidate_pool_requested", meta.get("n_candidates_requested", 50000)),
        ("unique_candidates_after_deduplication", filter_report.get("n_in", 49969)),
        ("candidates_after_feasibility_filtering", "not separately recorded"),
        ("candidates_after_out_of_distribution_control", filter_report.get("n_out", 46136)),
        ("final_scored_candidates", meta.get("n_candidates_after_dedup_and_filters", 46136)),
        ("top_candidates_exported", 200),
        ("primary_model", "CatBoost"),
        ("screening_score", "predicted high-delivery probability"),
        ("numeric_ood_control_rule", "0.01 and 0.99 quantiles of the reference training distribution for Size, Zeta Potential, and Admin."),
    ]
    return pd.DataFrame(rows, columns=["item", "value"])


def style_workbook(writer: pd.ExcelWriter) -> None:
    fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    side = Side(style="thin", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    body = Font(name="Times New Roman", size=10)
    header = Font(name="Times New Roman", size=10, bold=True)
    alignment = Alignment(vertical="center", wrap_text=True)
    for worksheet in writer.book.worksheets:
        worksheet.freeze_panes = "A2"
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = header if cell.row == 1 else body
                cell.fill = fill if cell.row == 1 else PatternFill(fill_type=None)
                cell.border = border
                cell.alignment = alignment
        for column in worksheet.columns:
            width = max(len(str(cell.value or "")) for cell in column)
            worksheet.column_dimensions[column[0].column_letter].width = min(max(width + 2, 8), 45)


def main() -> None:
    SUPPLEMENTARY_DIR.mkdir(parents=True, exist_ok=True)
    meta_path = SCREENING_DIR / "generation_meta.json"
    meta = json.loads(meta_path.read_text(encoding="utf-8")) if meta_path.exists() else {}
    sheets = {
        "screening_summary": screening_summary(meta),
        "generation_meta": item_value_json(meta_path) if meta_path.exists() else pd.DataFrame(),
        "all_scored_candidates": read_optional_csv("generated_candidates_scored.csv"),
        "top_200_candidates": read_optional_csv("paper_candidate_table_top200.csv"),
        "top_20_candidates": read_optional_csv("paper_candidate_table_top20.csv"),
        "top_10_candidates": read_optional_csv("paper_candidate_table_top10.csv"),
        "ct_specific_candidates": read_optional_csv("paper_candidate_table_CT_Breast_top20.csv"),
        "ct_specific_ranges": read_optional_csv("range_recommendation_CT_Breast.csv"),
        "range_detail": read_optional_csv("range_recommendation_detail.csv"),
        "range_pretty": read_optional_csv("range_recommendation_pretty.csv"),
        "range_subset": read_optional_csv("range_recommendation_subset.csv"),
    }
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        for sheet_name, frame in sheets.items():
            if not frame.empty:
                frame.to_excel(writer, sheet_name=sheet_name, index=False)
        range_meta = SCREENING_DIR / "range_recommendation_meta.json"
        if range_meta.exists():
            item_value_json(range_meta).to_excel(writer, sheet_name="range_meta", index=False)
        style_workbook(writer)
    print(f"Generated: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
