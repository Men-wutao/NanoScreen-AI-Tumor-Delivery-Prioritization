"""Build Supplementary Excel File B from repository-relative model outputs."""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path(__file__).resolve().parents[1]
RESULTS_DIR = PROJECT_ROOT / "results"
MODEL_SUMMARY_DIR = RESULTS_DIR / "model_summary"
SUPPLEMENTARY_DIR = PROJECT_ROOT / "supplementary"
OUTPUT_FILE = SUPPLEMENTARY_DIR / "Supplementary_Excel_File_B_model_evaluation_outputs.xlsx"


SUMMARY_FILES = [
    ("cv_metrics_raw", MODEL_SUMMARY_DIR / "model_metrics_summary_cv_raw.csv"),
    ("cv_metrics_display", MODEL_SUMMARY_DIR / "model_metrics_summary_cv_display.csv"),
    ("oof_metrics_summary", MODEL_SUMMARY_DIR / "model_metrics_summary_oof.csv"),
    ("test_metrics_summary", MODEL_SUMMARY_DIR / "model_metrics_summary_test.csv"),
    ("metrics_long", MODEL_SUMMARY_DIR / "model_metrics_summary_long.csv"),
    ("primary_model_cv_raw", MODEL_SUMMARY_DIR / "selected_primary_model_cv_raw.csv"),
    ("primary_model_cv_display", MODEL_SUMMARY_DIR / "selected_primary_model_cv_display.csv"),
    ("primary_model_oof", MODEL_SUMMARY_DIR / "selected_primary_model_oof.csv"),
    ("primary_model_test", MODEL_SUMMARY_DIR / "selected_primary_model_test.csv"),
]


def value_to_text(value: Any) -> Any:
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def model_name_from_file(path: Path, suffix: str) -> str:
    return path.name[: -len(suffix)] if path.name.endswith(suffix) else path.stem


def merge_csv_pattern(pattern: str, suffix: str) -> pd.DataFrame:
    rows = []
    for path in sorted(RESULTS_DIR.glob(pattern)):
        frame = pd.read_csv(path)
        frame["model_name"] = model_name_from_file(path, suffix)
        frame["source_file"] = path.name
        rows.append(frame)
    return pd.concat(rows, ignore_index=True, sort=False) if rows else pd.DataFrame()


def merge_reports() -> pd.DataFrame:
    rows = []
    for path in sorted(RESULTS_DIR.glob("*_report.json")):
        with path.open("r", encoding="utf-8") as handle:
            payload = json.load(handle)
        rows.append(
            {
                "model_name": payload.get("model_name", model_name_from_file(path, "_report.json")),
                "model_family": payload.get("model_family", ""),
                "best_cv_pr_auc": payload.get("best_cv_pr_auc", ""),
                "best_threshold": payload.get("best_threshold", ""),
                "best_oof_f1": payload.get("best_oof_f1", ""),
                "imbalance_method_run": payload.get("imbalance_method_run", ""),
                "best_params": value_to_text(payload.get("best_params", "")),
                "oof_metrics": value_to_text(payload.get("oof_metrics", "")),
                "test_metrics": value_to_text(payload.get("test_metrics", "")),
                "cv_summary": value_to_text(payload.get("cv_summary", "")),
                "source_file": path.name,
            }
        )
    return pd.DataFrame(rows)


def style_workbook(writer: pd.ExcelWriter) -> None:
    fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    side = Side(style="thin", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    body = Font(name="Times New Roman", size=10)
    header = Font(name="Times New Roman", size=10, bold=True)
    alignment = Alignment(vertical="center", wrap_text=True)
    for worksheet in writer.book.worksheets:
        worksheet.freeze_panes = "A2"
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = header if cell.row == 1 else body
                cell.fill = fill if cell.row == 1 else PatternFill(fill_type=None)
                cell.border = border
                cell.alignment = alignment
        for column in worksheet.columns:
            width = max(len(str(cell.value or "")) for cell in column)
            worksheet.column_dimensions[column[0].column_letter].width = min(max(width + 2, 8), 45)


def main() -> None:
    SUPPLEMENTARY_DIR.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        for sheet_name, path in SUMMARY_FILES:
            if path.exists():
                pd.read_csv(path).to_excel(writer, sheet_name=sheet_name, index=False)
        optional_sheets = {
            "optuna_trials_all": merge_csv_pattern("*_optuna_trials.csv", "_optuna_trials.csv"),
            "cv_fold_metrics_all": merge_csv_pattern("*_cv_fold_metrics.csv", "_cv_fold_metrics.csv"),
            "individual_oof_metrics": merge_csv_pattern("*_metrics_oof.csv", "_metrics_oof.csv"),
            "individual_test_metrics": merge_csv_pattern("*_metrics_test.csv", "_metrics_test.csv"),
            "oof_predictions_all": merge_csv_pattern("*_oof_predictions.csv", "_oof_predictions.csv"),
            "test_predictions_all": merge_csv_pattern("*_test_predictions.csv", "_test_predictions.csv"),
            "model_reports": merge_reports(),
        }
        for sheet_name, frame in optional_sheets.items():
            if not frame.empty:
                frame.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        style_workbook(writer)
    print(f"Generated: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
