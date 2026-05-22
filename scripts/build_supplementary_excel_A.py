"""Build Supplementary Excel File A from repository-relative inputs.

This script assembles dataset and label supplementary sheets. It expects the
curated dataset files to be available under data/ and writes the workbook to
the supplementary/ folder.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = PROJECT_ROOT / "data"
SUPPLEMENTARY_DIR = PROJECT_ROOT / "supplementary"
OUTPUT_FILE = SUPPLEMENTARY_DIR / "Supplementary_Excel_File_A_dataset_and_labels.xlsx"


SHEET_SOURCES = [
    ("cleaned_analytical_dataset", DATA_DIR / "processed_dataset" / "cleaned_analytical_dataset.csv", "csv"),
    ("train_dataset_with_label", DATA_DIR / "processed_dataset" / "train_dataset_with_label.csv", "csv"),
    ("test_dataset_with_label", DATA_DIR / "processed_dataset" / "test_dataset_with_label.csv", "csv"),
    ("dataset_with_split_label", DATA_DIR / "processed_dataset" / "analytical_dataset_with_split_and_label.csv", "csv"),
    ("endpoint_definition", DATA_DIR / "data_dictionary" / "endpoint_and_label_definition.json", "item_value_json"),
    ("variable_dictionary", DATA_DIR / "data_dictionary" / "variable_dictionary.csv", "csv"),
    ("threshold_sensitivity", DATA_DIR / "data_dictionary" / "threshold_sensitivity_report.csv", "csv"),
    ("variable_ranges_full", DATA_DIR / "data_dictionary" / "analytical_variable_ranges_full.csv", "csv"),
    ("variable_ranges_train", DATA_DIR / "data_dictionary" / "analytical_variable_ranges_train_only.csv", "csv"),
    ("category_levels", DATA_DIR / "data_dictionary" / "category_levels_summary.json", "category_json"),
]


def value_to_text(value: Any) -> Any:
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def read_source(path: Path, source_type: str) -> pd.DataFrame:
    if source_type == "csv":
        return pd.read_csv(path)
    with path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)
    if source_type == "item_value_json":
        return pd.DataFrame([{"item": key, "value": value_to_text(value)} for key, value in payload.items()])
    if source_type == "category_json":
        rows = []
        for variable, detail in payload.items():
            levels = detail.get("levels", []) if isinstance(detail, dict) else detail
            rows.append(
                {
                    "variable": variable,
                    "n_levels": detail.get("n_levels", len(levels)) if isinstance(detail, dict) else len(levels),
                    "levels": "; ".join(map(str, levels)) if isinstance(levels, list) else value_to_text(levels),
                }
            )
        return pd.DataFrame(rows)
    raise ValueError(f"Unsupported source type: {source_type}")


def style_workbook(writer: pd.ExcelWriter) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    body_font = Font(name="Times New Roman", size=10)
    header_font = Font(name="Times New Roman", size=10, bold=True)
    alignment = Alignment(vertical="center", wrap_text=True)
    for worksheet in writer.book.worksheets:
        worksheet.freeze_panes = "A2"
        for row in worksheet.iter_rows():
            for cell in row:
                cell.font = header_font if cell.row == 1 else body_font
                cell.fill = header_fill if cell.row == 1 else PatternFill(fill_type=None)
                cell.border = border
                cell.alignment = alignment
        for column in worksheet.columns:
            width = max(len(str(cell.value or "")) for cell in column)
            worksheet.column_dimensions[column[0].column_letter].width = min(max(width + 2, 8), 45)


def main() -> None:
    SUPPLEMENTARY_DIR.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        for sheet_name, path, source_type in SHEET_SOURCES:
            if not path.exists():
                print(f"WARNING: Missing input for sheet {sheet_name}: {path.name}")
                continue
            read_source(path, source_type).to_excel(writer, sheet_name=sheet_name[:31], index=False)
        style_workbook(writer)
    print(f"Generated: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
